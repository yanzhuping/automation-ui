import sys
from libs.test_utils import get_opt, get_global_config,readDataFromMySQL,deleteDataFromMySQL,get_root_path
from libs.interface_fun import createSession
from libs.interface_fun import get_value_from_json,dict_get
import datetime
import time
import random
from openpyxl import load_workbook,workbook
import os

class GetInformation():
    '''
    处理身份证号码，从中获取性别，年龄
    '''
    def __init__(self, id):
        self.id = id
        self.birth_year = int(self.id[6:10])
        self.birth_month = int(self.id[10:12])
        self.birth_day = int(self.id[12:14])

    def get_birthday(self):
        """通过身份证号获取出生日期"""
        birthday = "{0}-{1}-{2}".format(self.birth_year, self.birth_month, self.birth_day)
        return birthday

    def get_sex(self):
        """男生：1 女生：0"""
        num = int(self.id[16:17])
        if num % 2 == 0:
            return 0
        else:
            return 1

    def get_age(self):
        """通过身份证号获取年龄"""
        now = (datetime.datetime.now() + datetime.timedelta(days=1))
        year = now.year
        month = now.month
        day = now.day

        if year == self.birth_year:
            return 0
        else:
            if self.birth_month > month or (self.birth_month == month and self.birth_day > day):
                return year - self.birth_year - 1
            else:
                return year - self.birth_year

def timeStamp(timeNum):
    #将毫秒级的时间戳转换为格式化时间
    timeStamp = float(timeNum / 1000)
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d", timeArray)
    return otherStyleTime

def readAndWeiteXLSX(filepath,jsonkey,jsonvalue):
    '''
    操纵xlsx文档
    :param filepath: 文件路径
    :return:
    '''
    data_list = []
    wb =load_workbook(filepath)
    sheet = wb['sheet']
    #获取表格第一行的所有数据（行列的起始都是1开始算）
    for row in sheet[1]:
        data_list.append(row.value)
    index = data_list.index(jsonkey)
    sheet.cell(2,index+1).value = jsonvalue
    wb.save(filepath)
    wb.close()

def getRandomSet(bits=26):
    #随机生成一个26位的数字字母混合字符穿
    num_set = [chr(i) for i in range(48, 58)]
    char_set = [chr(i) for i in range(97, 123)]
    total_set = num_set + char_set
    value_set = "".join(random.sample(total_set, bits))
    return value_set

class CreateData():
    def __init__(self):
        # 存储一系列接种的返回的变量
        self.dict = {}
        self.now = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        self.t_opt = get_opt(sys.argv[1:])
        self.g_config = get_global_config(self.t_opt.get('env'))
        self.clientCode = eval(self.g_config.get('customer')).get(self.t_opt.get('customer'))
        self.s = createSession(
            self.g_config.get('username'),
            self.g_config.get('password'),
            self.clientCode)

    def getInfoFromDB(self):
        #从数据库获取姓名、身份证号码、电话号码、银行卡账号、员工编号
        g_config = self.g_config
        emName = readDataFromMySQL(g_config, 'tablename1', 'name')
        deleteDataFromMySQL(g_config, 'tablename1', 'name')
        idNo = readDataFromMySQL(g_config, 'tablename2', 'idNumber')
        deleteDataFromMySQL(g_config, 'tablename2', 'idNumber')
        bankNo = readDataFromMySQL(g_config, 'tablename3', 'bankCardNo')
        deleteDataFromMySQL(g_config, 'tablename3', 'bankCardNo')
        phoneNo = readDataFromMySQL(g_config, 'tablename4', 'phoneNo')
        deleteDataFromMySQL(g_config, 'tablename4', 'phoneNo')
        emNo = readDataFromMySQL(g_config, 'tablename5', 'emNo')
        deleteDataFromMySQL(g_config, 'tablename5', 'emNo')
        self.dict['emName'] = emName
        self.dict['idNo'] = idNo
        self.dict['bankNo'] = bankNo
        self.dict['phoneNo'] = phoneNo
        self.dict['emNo'] = emNo

    def getJobCode(self):
        #获取岗位编码,返回其中1个编码
        s = self.s
        url = f'https://qa.tranderpay.com/api/v1/orgPosition/selectOrgPosition?clientCode={self.clientCode}&currPage=1&pageSize=18'
        result = s.request('get', url, verify=False, timeout=20)
        # print(result.json())
        temlist = get_value_from_json('positionCode',result.json(),[])[0]
        # print(temlist)
        return temlist

    def getPayrollGroupsID(self):
        #获取薪资组ID
        s = self.s
        url = r'https://qa.tranderpay.com/api/v1/clientPayrollGroups/list'
        result = s.request('get', url, verify=False, timeout=20)
        # print(result.json())
        cpgId = get_value_from_json('cpgId',result.json(),[])[0]
        payrollGroupCn = get_value_from_json('payrollGroupCn',result.json(),[])[0]
        # print(cpgId)
        self.dict['cpgId'] = cpgId
        self.dict['payrollGroupCn'] = payrollGroupCn
        return cpgId

    def getJobInof(self):
        #查询所有得职位岗位部门晋升通道职级
        s = self.s
        jobCode = self.getJobCode()
        url = f'https://qa.tranderpay.com/api/v1/orgPosition/selectByOpsitionId?clientCode={self.clientCode}&opsitionCode={jobCode}'
        # print(url)
        result = s.request('get',url,verify=False, timeout=20)
        # print('result:',result.json())
        legalEntityCode = get_value_from_json('legalId',result.json(),[])[0]
        jobCode = get_value_from_json('jobCode',result.json(),[])[0]
        jobDesc = get_value_from_json('jobDesc',result.json(),[])[0]
        organizationCode = get_value_from_json('organizationCode',result.json(),[])[0]
        organizationDesc = get_value_from_json('organizationDesc',result.json(),[])[0]
        positionCode = get_value_from_json('positionCode',result.json(),[])[0]
        positionName = get_value_from_json('positionName',result.json(),[])[0]
        #查询所有的职位岗位部门晋升通道职级
        url1 = f'https://qa.tranderpay.com/api/v1/platformLegalServicesItems/selectServicesItemsByCode?legalEntityCode={legalEntityCode}&clientCode={self.clientCode}'
        result1 = s.request('get', url1, verify=False, timeout=20)
        # print('result1:',result1.json())
        clientLegalCode = get_value_from_json('clientLegalCode',result1.json(),[])[0]
        legalEntityDescription = get_value_from_json('legalEntityDescription',result1.json(),[])[0]
        #成本中心分页查询
        url2 = f'https://qa.tranderpay.com/api/v1/costCenter/listByPage?currPage=1&pageSize=1000&leName={clientLegalCode}'
        result2 = s.request('get', url2, verify=False, timeout=20)
        # print('result2:',result2.json())
        self.dict['jobCode'] = jobCode
        self.dict['jobDesc'] = jobDesc
        self.dict['organizationCode'] = organizationCode
        self.dict['organizationDesc'] = organizationDesc
        self.dict['positionCode'] = positionCode
        self.dict['positionName'] = positionName
        self.dict['clientLegalCode'] = clientLegalCode
        self.dict['legalEntityDescription'] = legalEntityDescription
        return jobCode,jobDesc,organizationCode,organizationDesc,positionCode,positionName,clientLegalCode,legalEntityDescription

    def getStaffPayrollData(self):
        #查询员工薪资数据
        s =self.s
        url = f'https://qa.tranderpay.com/api/v1/clientPayrollGroups/searchPayrollGroup?cpgId={self.getPayrollGroupsID()}'
        result = s.request('get', url, verify=False, timeout=20)
        # print(result.json())
        #时间戳需要转换‘2021-12-01’
        payDate = timeStamp(get_value_from_json('payDate',result.json(),[])[0])
        payrollBelongEnd = timeStamp(get_value_from_json('payrollBelongEnd',result.json(),[])[0])
        payrollBelongMonth = timeStamp(get_value_from_json('payrollBelongMonth',result.json(),[])[0])
        payrollBelongStart = timeStamp(get_value_from_json('payrollBelongStart',result.json(),[])[0])
        #当前日期的毫秒级时间戳
        payrollBelongStart_1 = get_value_from_json('payrollBelongStart',result.json(),[])[0]
        #'202112'需要转换为‘2021-12’
        iitMonth = get_value_from_json('iitMonth',result.json(),[])
        iitMonth = "{0}-{1}".format(iitMonth[0:4], iitMonth[5:7])
        if iitMonth[5:7] == '12':
            iitMonth_1 = "{0}-{1}".format(int(iitMonth[0:4])+1, 1)
        else:
            iitMonth_1 = "{0}-{1}".format(iitMonth[0:4], int(iitMonth[5:7])+1)
        #获取薪资组固定薪资项
        url1 = f'https://qa.tranderpay.com/api/v1/clientPayrollGroups/searchPayrollItem?employeeCode=&cpgId={self.getPayrollGroupsID()}'
        result1 = s.request('get', url1, verify=False, timeout=20)
        # print(result1.json())
        payrollItem = get_value_from_json('data',result1.json(),[])[0]
        self.dict['payrollItem'] = payrollItem
        self.dict['payDate'] = payDate
        self.dict['payrollBelongEnd'] = payrollBelongEnd
        self.dict['payrollBelongMonth'] = payrollBelongMonth
        self.dict['payrollBelongStart'] = payrollBelongStart
        self.dict['payrollBelongStart_1'] = payrollBelongStart_1
        self.dict['iitMonth'] = iitMonth
        self.dict['iitMonth_1'] = iitMonth_1
        return payrollItem,payDate,payrollBelongEnd,payrollBelongMonth,payrollBelongStart,iitMonth,iitMonth_1

    def createDataFromStaffManage(self):
        #从员工管理-员工数据维护添加1条数据
        s = self.s
        url = r'https://qa.tranderpay.com/api/v1/employee/updateByAll'
        false = False
        true = True
        null = None
        body = {
            "eeName": self.dict.get('emName'),
            "idNumber": self.dict.get('idNo'),
            "idType": "1",
            "nationality": "CN",
            "gender": GetInformation(self.dict.get('idNo')).get_sex(),
            "dateOfBirth": GetInformation(self.dict.get('idNo')).get_birthday(),
            "isHandicappe": false,
            "certHandicappe": "",
            "contactsMnccode": "86",
            "contactsPersonalEmail": "",
            "contactsMobile": self.dict.get('phoneNo'),
            "addressDetails": [{
                "contactsAddressType": 0,
                "contactsAddressProvince": "310000",
                "contactsAddressProvinceCn": "上海市",
                "contactsAddressCity": "310100",
                "contactsAddressCityCn": "上海市",
                "contactsAddressDistrict": "",
                "contactsAddressDistrictCn": "",
                "contactsAddressDetails": "四川北路1314号",
                "addressSame": null,
                "addressSameOneActive": false,
                "addressSameTwoActive": false
            }],
            "legalEntityDesc": self.dict.get('legalEntityDescription'),
            "employeeCode": "",
            "erEeCode": self.dict.get('emNo'),
            "employmentType": 1,
            "employeeType": 1,
            "employeeStatus": 2,
            "legalEntityCode": self.dict.get('clientLegalCode'),
            "deptCode": self.dict.get('organizationCode'),
            "deptDesc": self.dict.get('organizationDesc'),
            "position": self.dict.get('positionName'),
            "positionCode": self.dict.get('positionCode'),
            "job": self.dict.get('jobDesc'),
            "jobCode": self.dict.get('jobCode'),
            "bizEmail": "",
            "firstHiringDate": "",
            "firstDateOfJoin": self.now,
            "dateOfJoin": self.now,
            "probationEndDate": "2022-06-01",
            "intenshipEndDate": "",
            "terminationDate": "",
            "contractPeriodStart": self.now,
            "contractPeriodEnd": "9999-12-31",
            "contractPeriodData": [self.now, "9999-12-31"],
            "workProvince": "310000",
            "workCity": "310100",
            "workDistrict": "",
            "areaName": "",
            "payrollGroup": self.dict.get('cpgId'),
            "payrollBelongMonth": self.dict.get('payrollBelongMonth'),
            "payDate": self.dict.get('payDate'),
            "payslipDate": "",
            "payrollDate": "",
            "payrollAccountTo": "",
            "payrollCycle": [self.dict.get('payrollBelongStart'), self.dict.get('payrollBelongEnd')],
            "payrollBelongStart": self.dict.get('payrollBelongStart'),
            "payrollBelongEnd": self.dict.get('payrollBelongEnd'),
            "bankAccountOne": self.dict.get('bankNo'),
            "bankNameOne": "测试银行账号",
            "payeeNameOne": self.dict.get('emName'),
            "bankPercentageOne": 100,
            "bankAccountTwo": "",
            "bankNameTwo": "",
            "payeeNameTwo": "",
            "bankPercentageTwo": "",
            "employeeBankDetails": [{
                "bankAccount": self.dict.get('bankNo'),
                "copyValue": "",
                "bankName": "测试银行账号",
                "payeeName": self.dict.get('emName'),
                "bankPercentage": 100,
                "showValue": true,
                "key": self.dict.get('payrollBelongStart_1')
            }],
            "hukouType": "",
            "socialBenefitsProvince": "310000",
            "socialBenefitsProvinceName": "",
            "socialBenefitsCity": "310100",
            "socialBenefitsCityName": "",
            "socialBenefitsDistrict": "",
            "socialBenefitsAreaName": "",
            "socialBenefitsBase": "",
            "socialBenefitsAccount": "",
            "sbStartMonth": "",
            "sbEndMonth": "",
            "housingFundProvince": "310000",
            "housingFundProvinceName": "",
            "housingFundCity": "310100",
            "housingFundCityName": "",
            "housingFundDistrict": "",
            "housingFundAreaName": "",
            "housingFundBase": "",
            "housingFundAccount": "",
            "hfStartMonth": "",
            "hfEndMonth": "",
            "medicalBase": "",
            "taxType": "",
            "taxProvince": "310000",
            "taxCity": "310100",
            "taxDistrict": "",
            "entryYearEmploymentSituation": "",
            "taxServingType": 1,
            "taxContributionMonthStr1": self.dict.get('iitMonth_1'),
            "taxBelongMonth": "",
            "iitMonth": self.dict.get('iitMonth'),
            "taxCompanyName": self.dict.get('clientLegalCode'),
            "taxCompanyNameCn": "",
            "taxStatus": "",
            "taxServingDate": self.now,
            "taxHireHouse": "",
            "taxFirstEntryDate": "",
            "taxDepartureDate": "",
            "companyYos": "",
            "socialityYos": "",
            "payrollBeginDate": self.now,
            "lastWorkingDay": "",
            "terminationReason": "",
            "employeeCostCenterDatas": [],
            "ccCodeA": "",
            "percentageA": "",
            "beginDateA": "",
            "endDateA": "",
            "ccCodeB": "",
            "percentageB": "",
            "beginDateB": "",
            "endDateB": "",
            "costCenterCode": "",
            "payrollItem": self.dict.get('payrollItem'),
            "taxContributionMonth": ""
        }
        result = s.request('post', url, json=body, verify=False, timeout=20)
        print(result.json())
        print('创建的员工姓名：',self.dict.get('emName'))

    def findFieldMenuList(self,locators):
        #人事管理员工分类查询
        s = self.s
        url = f'https://qa.tranderpay.com/api/v1/personnelManagement/findFieldMenuList?clientCode={self.clientCode}'
        result = s.request('get', url, verify=False, timeout=20)
        templateMenuId = dict_get(result.json(),locators)
        print(templateMenuId)
        return templateMenuId

    def addFieldValeByEmployee(self, locators, temporaryCode):
        #员工信息分类临时数据查询redis/mogo
        #页面的小保存
        s = self.s
        url = f'https://qa.tranderpay.com/api/v1/personnelManagement/findTemporaryFieldByMenuIdDbOrRedis?templateMenuId={self.findFieldMenuList(locators)}&clientCode={self.clientCode}&temporaryCode={temporaryCode}'
        result = s.request('get', url, verify=False, timeout=20)
        menuFieldRelList = get_value_from_json('data', result.json(),[])[0]
        url1 = r'https://qa.tranderpay.com/api/v1/personnelManagement/addFieldValeByEmployee'
        body = {
            "templateMenuId": self.findFieldMenuList(locators),
            "temporaryCode": temporaryCode,
            "menuFieldRelList": menuFieldRelList,
            "clientCode": self.clientCode
        }
        return menuFieldRelList
        #将uuid写在配置文件中比较方便，明天操作、、、

    def readAndWeiteXLSXPlus(self,filepath, **kwargs):
        '''
        操纵xlsx文档
        :param filepath: 文件路径
        :return:
        '''
        data_list = []
        wb = load_workbook(filepath)
        sheet = wb['sheet']
        # 获取表格第一行的所有数据（行列的起始都是1开始算）
        for row in sheet[1]:
            data_list.append(row.value)
        # print(kwargs)
        for key in kwargs.keys():
            # print(key)
            index = data_list.index(key)
            # print(index)
            # print(kwargs[key])
            sheet.cell(2, index + 1).value = kwargs[key]
        wb.save(filepath)
        wb.close()

    def uploadExcel(self,fileName):
        s = self.s
        url = r'https://qa.tranderpay.com/api/v1/uploadFile/uploadExcel'
        body = {'clientCode':self.clientCode,'importType':'employeeManagementImportEntry'}
        files = {'file':(fileName,
                         open(os.path.join(get_root_path(),
                                           'test_data',
                                           '人事管理之员工信息',
                                           fileName
                                           ),'rb'
                              )
                         )
                 }
        result = s.request('post', url=url, data=body, files=files, verify=False,timeout=20)
        # print(result.json())
        message = get_value_from_json('message',result.json(),[])[0]
        url1 = r'https://qa.tranderpay.com/api/v1/employeePersonnelImport/importEmployeeExcel'
        body1 = {"queryType":1,
                 "url":message,
                 "importType":"employeeManagementImportEntry"
                 }
        print(body1)
        result1 = s.request('post', url=url1, json=body1, verify=False, timeout=20)
        print(result1)
        print(result1.json())

    def createDataFromPersonnelManage(self,fileName):
        #从人事管理-员工信息添加1条数据
        filepath = os.path.join(get_root_path(), 'test_data', '人事管理之员工信息', fileName)
        sex = GetInformation(self.dict.get('idNo')).get_sex()
        if sex == 0:
            sex = '女'
        if sex == 1:
            sex = '男'
        birthday = GetInformation(self.dict.get('idNo')).get_birthday().replace('-','/')

        age = GetInformation(self.dict.get('idNo')).get_age()
        self.readAndWeiteXLSXPlus(filepath,
                                  证件号码=self.dict.get('idNo'),
                                  姓名=self.dict.get('emName'),
                                  性别=sex,
                                  出生日期=birthday,
                                  年龄=age,
                                  手机号=self.dict.get('phoneNo'),
                                  员工企业编号=self.dict.get('emNo'),
                                  入职日期=self.now.replace('-','/'),
                                  岗位=self.dict.get('positionName'),
                                  岗位编码=self.dict.get('positionCode'),
                                  职位=self.dict.get('jobDesc'),
                                  职位编码=self.dict.get('jobCode'),
                                  部门=self.dict.get('organizationDesc'),
                                  部门编码=self.dict.get('organizationCode'),
                                  所属法人实体=self.dict.get('legalEntityDescription'),
                                  所属法人实体编码=self.dict.get('clientLegalCode'),
                                  开户银行账号=self.dict.get('bankNo'),
                                  开户名=self.dict.get('emName'),
                                  薪资组=self.dict.get('payrollGroupCn'),
                                  计薪开始日期=self.now.replace('-','/'),
                                  发薪日期=self.dict.get('payDate').replace('-','/'),
                                  )
        self.uploadExcel(fileName)
        print('创建的员工姓名：', self.dict.get('emName'))

def run(func):
    start = datetime.datetime.now()
    a = CreateData()

    if func == '员工管理添加一条数据':
        a.getInfoFromDB()
        a.getJobCode()
        a.getPayrollGroupsID()
        a.getJobInof()
        a.getStaffPayrollData()
        a.createDataFromStaffManage()

    if func == '人事管理添加一条数据':
        fileName = '上海自动化测试有限公司创建数据模板_1.xlsx'
        a.getInfoFromDB()
        a.getJobCode()
        a.getJobInof()
        a.getPayrollGroupsID()
        a.getStaffPayrollData()
        a.createDataFromPersonnelManage(fileName)

    end = datetime.datetime.now()
    print("程序运行时间：", end - start)

if __name__ == '__main__':
    # python create_main.py --env=rszy --customer=上海自动化测试有限公司 --func=人事管理添加一条数据
    filepath = r'D:\WorkCode\automation\test_data\人事管理之员工信息\上海自动化测试有限公司创建数据模板.xlsx'
    # readAndWeiteXLSXPlus(filepath, 手机号='13088888888',政治面貌='党员')
